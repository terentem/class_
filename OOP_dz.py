from openpyxl import Workbook
import xlrd
import json

class F:
    def __init__(self,list_):
        self.list_=list_
        self.dict_=None
        self.file_name=None
        self.new_dict=None
        self.w_sh=None
        self.dict_from_xls=None

    def convert_list_to_dict(self):
        self.dict_={}
        for x in self.list_:
            self.dict_[x]=x
        print('dict_',self.dict_)

    def save_file(self,file_name):
        with open(file_name, "w") as write_file:
               json.dump(self.dict_, write_file)
        write_file.close()

    def read_file(self,file_name):
        with open(file_name, "r") as read_file:
            self.new_dict = json.load(read_file)
        read_file.close()
        return self.new_dict

    def export_dict_to_xls(self, file_name):
        wb=Workbook()
        self.w_sh=wb.active
        i=1
        self.w_sh.cell(1,1).value='Key'
        self.w_sh.cell(1,2).value='Value'
        for k,v in self.new_dict.items():
            i+=1
            self.w_sh.cell(i,1).value=v
            self.w_sh.cell(i,2).value = v
        wb.save(file_name)
        wb.close()


    def import_from_xls(self,file_name):
        wb = xlrd.open_workbook(file_name)
        sh = wb.sheet_by_index(0)
        p=sh.nrows
        z=1
        self.dict_from_xls={}
        while z<p-1:
             key=sh.cell(z,0).value
             val=sh.cell(z,1).value
             self.dict_from_xls[key]=val
             z+=1
        return self.dict_from_xls

#0.Начальные данные
list_=[11,22,33,44,55]
class_F=F(list_)

#1. Конвертируем лист в дикт
class_F.convert_list_to_dict()

#2.Сохраняем дикт в файл
class_F.save_file('file_dz')

#3. Вычитываем дикт из файла
new_dict=class_F.read_file('file_dz')
print('new_dict=',new_dict,'type_new_dict',type(new_dict))

#4. Экспортируем дикт в xls
class_F.export_dict_to_xls('DICT.xlsx')

#5. СОздаём дикт из данных xls файла
dict_from_xls=class_F.import_from_xls('DICT.xlsx')
print('dict_from_xls',type(dict_from_xls),dict_from_xls)